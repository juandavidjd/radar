#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ciclo_fecha.py
Fecha -> posiciones -> 50 combinaciones -> asigna a TODAS las figuras de la plantilla (hasta 38)
y deja el resto sueltas (con 38 figuras, quedan 12 sueltas).

Aleatoriedad:
- Se barajan las 50 combinaciones.
- Se barajan las figuras detectadas en la plantilla.
- Se emparejan combos con figuras al azar (no 1:1 secuencial).

Archivos:
- todo_pos.csv / .xlsx
- todo_sum.csv / .xlsx
- todo_fig.csv / .xlsx
- combos_sueltos.csv / .xlsx
- todo_map.csv / .xlsx

Bloqueos fijos:
- Fig 21 → 2x2 en C–D
- Fig 22 → 2x2 en D–E
- Fig 23 → 2x2 en E–F
"""

import csv, re, sys, random, argparse
from itertools import combinations
from datetime import datetime
from pathlib import Path

# Pandas opcional (para XLSX). Si falta, se generan igual todos los CSV.
try:
    import pandas as pd  # type: ignore
except Exception:
    pd = None

PLANTILLAS = [Path("plantilla_fig.xlsx"), Path("plantilla_fig.csv")]
LOCKED_COLS = {21:(2,3), 22:(3,4), 23:(4,5)}  # C–D, D–E, E–F
COLS = ["figura","a","b","c","d","e","f"]
NUM_RE = re.compile(r"^\s*[-+]?\d+(?:[.,]0+)?\s*$")

# ---------- Utilidades de fecha / posiciones ----------
def simplificar(n:int)->int:
    while n >= 10:
        n = sum(int(d) for d in str(n))
    return n

def parse_fecha(s:str):
    s = s.strip().replace("-", "/").replace(" ", "")
    for fmt in ("%d/%m/%Y", "%d/%m/%y"):
        try:
            dt = datetime.strptime(s, fmt)
            y = dt.year if "%Y" in fmt else (2000 + dt.year % 100)
            return dt.day, dt.month, y
        except ValueError:
            pass
    raise ValueError("Formato inválido. Usa DD/MM/YY o DD/MM/YYYY (p.ej. 09/08/2025).")

def calcular_posiciones(fecha_str:str):
    dd, mm, yyyy = parse_fecha(fecha_str)
    d = f"{dd:02d}"; m = f"{mm:02d}"; a = f"{yyyy%100:02d}"
    d1, d2 = int(d[0]), int(d[1])
    m1, m2 = int(m[0]), int(m[1])
    a1, a2 = int(a[0]), int(a[1])
    spec = [
        ("1º", d1, m1, ""), ("2º", d1, m2, ""),
        ("3º", d1, "", a1), ("4º", d1, "", a2),
        ("5º", d2, m1, ""), ("6º", d2, m2, ""),
        ("7º", d2, "", a1), ("8º", d2, "", a2),
    ]
    filas, sumas = [], []
    fecha_norm = f"{dd:02d}/{mm:02d}/{yyyy}"
    for pos, x, ym, ya in spec:
        y = ym if ym != "" else ya
        s_cruda = int(x) + int(y)
        s = simplificar(s_cruda)
        filas.append({
            "posicion": pos, "dia": x,
            "mes": ym if ym != "" else "",
            "ano": ya if ya != "" else "",
            "suma_cruda": s_cruda, "suma": s,
            "fecha": fecha_norm
        })
        sumas.append(s)
    return filas, sumas

# ---------- Combinaciones ----------
def generar_combos(sumas8, n=50):
    """Toma las 8 cifras y genera n combinaciones (4 índices de 8) y las convierte a string."""
    idx = list(range(8))
    pool = list(combinations(idx, 4))  # 70 posibles
    sel = random.sample(pool, min(n, len(pool)))
    return ["".join(str(sumas8[i]) for i in c) for c in sel]

# ---------- Lectura de plantilla ----------
def leer_plantilla(path:Path):
    if not path.exists():
        raise FileNotFoundError(f"No existe {path}")
    rows = []
    if path.suffix.lower() == ".xlsx":
        if pd is None:
            raise RuntimeError("Instala pandas+openpyxl para leer .xlsx")
        df = pd.read_excel(path, dtype=str).fillna("")
        cols = list(df.columns)
        if cols:
            cols[0] = "figura"; df.columns = cols
        for c in COLS:
            if c not in df.columns: df[c] = ""
        df = df[COLS]
        rows = [[str(v) for v in r] for _, r in df.iterrows()]
    else:
        with path.open(newline="", encoding="utf-8-sig") as f:
            for r in csv.reader(f):
                rows.append(r)

    # agrupar por bloques (figura con filas subsiguientes sin id)
    fig_grids, fig = {}, None
    for r in rows:
        r = (r + [""] * 7)[:7]
        head = (r[0] or "").strip()
        a_f = [(r[i] or "").strip() for i in range(1, 7)]
        if head and head.isdigit():
            fig = int(head)
            fig_grids.setdefault(fig, [])
            fig_grids[fig].append(a_f)
        else:
            if fig is None:
                continue
            fig_grids[fig].append(a_f)
    return fig_grids

# ---------- Lógica de relleno ----------
def posiciones_huecos(grid):
    pos = []
    for i, row in enumerate(grid):
        for j, val in enumerate(row):
            if str(val).strip() != "":
                pos.append((i, j))
    return pos

def asegurar_altura(grid, h=2):
    out = [row[:] for row in grid]
    while len(out) < h:
        out.append([""] * 6)
    return out

def forzar_bloque(grid, c1, c2):
    out = asegurar_altura([(row + [""] * 6)[:6] for row in grid], 2)
    for i in range(len(out)):
        for j in range(6):
            out[i][j] = ""
    out[0][c1] = out[0][c2] = out[1][c1] = out[1][c2] = "*"
    return out

def rellenar_con_combo(grid, combo4):
    out = [row[:] for row in grid]
    coords = posiciones_huecos(out)
    if len(coords) < 4:
        out = forzar_bloque(out, 2, 3)
        coords = posiciones_huecos(out)
    coords = coords[:4]
    digs = list(combo4) + [""] * 4
    for k, (i, j) in enumerate(coords):
        out[i][j] = digs[k]
    return out

def rellenar_figura(fig, grid, combo4):
    if fig in LOCKED_COLS:
        c1, c2 = LOCKED_COLS[fig]
        base = forzar_bloque(grid, c1, c2)
        return rellenar_con_combo(base, combo4)
    else:
        holes = posiciones_huecos(grid)
        if len(holes) != 4:
            grid = forzar_bloque(grid, 2, 3)
        return rellenar_con_combo(grid, combo4)

# ---------- Guardado CSV ----------
def save_csv_dict(rows, header, path):
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=header)
        w.writeheader(); w.writerows(rows)

def save_csv_rows(rows, header, path):
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(header); w.writerows(rows)

# ---------- Guardado XLSX (openpyxl si disponible) ----------
def save_xlsx_dict(rows, header, path):
    if pd is None:
        print(f"⚠️ No se creó '{path}' (falta pandas).")
        return
    try:
        import openpyxl  # noqa: F401
    except Exception:
        print(f"⚠️ No se creó '{path}' (falta openpyxl).")
        return
    df = pd.DataFrame(rows, columns=header)
    with pd.ExcelWriter(path, engine="openpyxl") as w:
        sheet = Path(path).stem
        df.to_excel(w, index=False, sheet_name=sheet)
        try:
            ws = w.sheets[sheet]
            from openpyxl.utils import get_column_letter
            for i, c in enumerate(df.columns, start=1):
                maxlen = max(len(str(c)), *(len(str(x)) for x in df[c].head(500)))
                ws.column_dimensions[get_column_letter(i)].width = max(10, min(24, maxlen + 2))
        except Exception:
            pass

def save_xlsx_rows(rows, header, path):
    if pd is None:
        print(f"⚠️ No se creó '{path}' (falta pandas).")
        return
    try:
        import openpyxl  # noqa: F401
    except Exception:
        print(f"⚠️ No se creó '{path}' (falta openpyxl).")
        return
    df = pd.DataFrame(rows, columns=header)
    with pd.ExcelWriter(path, engine="openpyxl") as w:
        sheet = Path(path).stem
        df.to_excel(w, index=False, sheet_name=sheet)
        try:
            ws = w.sheets[sheet]
            from openpyxl.utils import get_column_letter
            for i, c in enumerate(df.columns, start=1):
                maxlen = max(len(str(c)), *(len(str(x)) for x in df[c].head(500)))
                ws.column_dimensions[get_column_letter(i)].width = max(10, min(24, maxlen + 2))
        except Exception:
            pass

def guardar_todo_fig_csv(fig_grids_filled, path="todo_fig.csv"):
    rows = []
    for fig in sorted(fig_grids_filled.keys()):
        g = fig_grids_filled[fig]
        rows.append([fig] + g[0])
        for r in g[1:]:
            rows.append([""] + r)
    save_csv_rows(rows, COLS, path)

def guardar_todo_fig_xlsx(fig_grids_filled, path="todo_fig.xlsx"):
    rows = []
    for fig in sorted(fig_grids_filled.keys()):
        g = fig_grids_filled[fig]
        rows.append([fig] + g[0])
        for r in g[1:]:
            rows.append([""] + r)
    save_xlsx_rows(rows, COLS, path)

# ---------- Main ----------
def main():
    ap = argparse.ArgumentParser(description="Ciclo completo: fecha -> posiciones -> combos -> figuras (hasta 38).")
    ap.add_argument("--seed", type=int, default=None,
                    help="Semilla para aleatoriedad (opcional, para reproducibilidad).")
    args = ap.parse_args()
    if args.seed is not None:
        random.seed(args.seed)

    try:
        fecha = input("Ingrese la fecha (DD/MM/YY o DD/MM/YYYY): ").strip()
        filas_pos, sumas8 = calcular_posiciones(fecha)
    except Exception as e:
        print(f"❌ Error de fecha: {e}")
        sys.exit(1)

    # 50 combos barajados
    combos50 = generar_combos(sumas8, n=50)
    random.shuffle(combos50)

    # Plantilla
    plantilla = next((p for p in PLANTILLAS if p.exists()), None)
    if plantilla is None:
        print("❌ No se encontró plantilla_fig.xlsx ni plantilla_fig.csv")
        sys.exit(1)

    try:
        fig_grids = leer_plantilla(plantilla)
    except Exception as e:
        print(f"❌ Error leyendo plantilla: {e}")
        sys.exit(1)

    # Figuras reales en plantilla (asignamos hasta 38)
    all_figs = sorted(fig_grids.keys())
    if not all_figs:
        print("❌ La plantilla no contiene figuras.")
        sys.exit(1)

    target_assign = min(38, len(all_figs))  # si hay 38, asignamos 38 → 12 sueltas
    combos_assign = combos50[:target_assign]
    combos_sueltos = combos50[target_assign:]

    # Elegir aleatoriamente qué figuras de la plantilla se asignan
    figs_assign = all_figs[:]
    random.shuffle(figs_assign)
    figs_assign = figs_assign[:target_assign]

    # Relleno aleatorio + mapa
    filled, fmap = {}, {}
    for fig, combo in zip(figs_assign, combos_assign):
        grid = fig_grids.get(fig, [[""]*6, [""]*6])
        filled[fig] = rellenar_figura(fig, grid, combo)
        fmap[fig] = combo

    # Preservar TODAS las figuras no asignadas (tal cual vienen)
    for fig in all_figs:
        if fig not in filled:
            filled[fig] = fig_grids[fig]

    # ---- Guardar POS ----
    save_csv_dict(filas_pos, ["posicion","dia","mes","ano","suma_cruda","suma","fecha"], "todo_pos.csv")
    save_xlsx_dict(filas_pos, ["posicion","dia","mes","ano","suma_cruda","suma","fecha"], "todo_pos.xlsx")

    # ---- Guardar SUM ----
    save_csv_rows([[c] for c in combos50], ["combinacion"], "todo_sum.csv")
    save_xlsx_rows([[c] for c in combos50], ["combinacion"], "todo_sum.xlsx")

    # ---- Guardar FIG ----
    guardar_todo_fig_csv(filled, "todo_fig.csv")
    guardar_todo_fig_xlsx(filled, "todo_fig.xlsx")

    # ---- Guardar sueltos ----
    save_csv_rows([[c] for c in combos_sueltos], ["combinacion"], "combos_sueltos.csv")
    save_xlsx_rows([[c] for c in combos_sueltos], ["combinacion"], "combos_sueltos.xlsx")

    # ---- Guardar mapa (CSV + XLSX) ----
    map_rows = [[k, v] for k, v in sorted(fmap.items())]
    save_csv_rows(map_rows, ["figura","combinacion"], "todo_map.csv")
    save_xlsx_rows(map_rows, ["figura","combinacion"], "todo_map.xlsx")

    # --------- Resumen por consola ----------
    print("\n✅ Proceso completo.")
    print(f"   8 cifras base (suma): {sumas8}")
    print(f"   Figuras detectadas en plantilla: {len(all_figs)}")
    print(f"   Asignadas aleatoriamente: {target_assign} figuras")
    print(f"   Combinaciones sueltas: {len(combos_sueltos)}  (con 38 figuras deben ser 12)")
    print("   Plantilla usada:", plantilla.name)
    print("   Archivos:")
    print("   - todo_pos.csv / .xlsx")
    print("   - todo_sum.csv / .xlsx")
    print("   - todo_fig.csv / .xlsx")
    print("   - combos_sueltos.csv / .xlsx")
    print("   - todo_map.csv / .xlsx")

    print("\n📋 Asignación figura ↔ combinación:")
    for fig in sorted(fmap):
        print(f"   Figura {fig:02d} → {fmap[fig]}")

    if combos_sueltos:
        print("\n🗂️  Combos sueltos:")
        for i, c in enumerate(combos_sueltos, 1):
            print(f"   {i:02d}. {c}")

    print("\nMuestra las figuras con:")
    print("  python leer_figuras.py --input todo_fig.xlsx")
    if args.seed is not None:
        print(f"\n(Reproducible: seed={args.seed})")

if __name__ == "__main__":
    main()
