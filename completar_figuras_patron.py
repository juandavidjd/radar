#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
completar_figuras_patron.py
Crea/Completa 'todo_fig.csv' y 'todo_fig.xlsx' con 100 figuras (a..f)
usando un patrón fijo de asteriscos:

Fila 0: C D E F = '*'
Fila 1: C = '*'
Fila 2: D = '*'
Fila 3: E = '*'
Fila 4: F = '*'
Fila 5: C y F = '*'
Fila 6: C y E = '*'

Es decir, 7 filas de alto x 6 columnas (a..f). Se preservan figuras que ya existan.
Si faltan, se agregan siguiendo este patrón. No coloca números, solo '*'.

Requiere pandas+openpyxl solo para escribir .xlsx. CSV se genera siempre.
"""

import csv
from pathlib import Path

try:
    import pandas as pd
except Exception:
    pd = None

OUT_CSV  = Path("todo_fig.csv")
OUT_XLSX = Path("todo_fig.xlsx")
COLS     = ["figura","a","b","c","d","e","f"]
TARGET   = 100

def read_existing():
    """Lee todo_fig.xlsx o todo_fig.csv si existen, y devuelve dict fig->grid."""
    rows = None
    if OUT_XLSX.exists() and pd is not None:
        try:
            df = pd.read_excel(OUT_XLSX, dtype=str).fillna("")
            cols = list(df.columns)
            if cols:
                cols[0]="figura"; df.columns = cols
            for c in COLS:
                if c not in df.columns: df[c]=""
            df = df[COLS]
            rows = [[str(v) for v in r] for _,r in df.iterrows()]
        except Exception:
            rows = None
    if rows is None and OUT_CSV.exists():
        rows=[]
        with OUT_CSV.open(newline="", encoding="utf-8-sig") as f:
            r = csv.reader(f)
            data = [row for row in r]
        if data:
            head = [h.strip().lower() for h in data[0]]
            data = data[1:] if head and "figura" in head else data
            for row in data:
                rows.append([c.strip() for c in (row + [""]*7)[:7]])
    # agrupar por figura
    figs = {}
    if rows:
        cur = None
        for r in rows:
            r = (r + [""]*7)[:7]
            head = (r[0] or "").strip()
            a_f  = [(r[i] or "").strip() for i in range(1,7)]
            if head.isdigit():
                cur = int(head)
                figs.setdefault(cur, [])
                figs[cur].append(a_f)
            else:
                if cur is None: continue
                figs[cur].append(a_f)
        # limpia colas vacías
        for k,g in list(figs.items()):
            while g and all(c=="" for c in g[-1]): g.pop()
            if not g: figs.pop(k,None)
    return figs

def pattern_grid():
    """
    Devuelve el patrón 7x6 (a..f) con '*' EXACTAMENTE como lo pediste.
    """
    g = [[""]*6 for _ in range(7)]
    # Helpers (índices 0-based): a=0,b=1,c=2,d=3,e=4,f=5
    C,D,E,F = 2,3,4,5
    # Fila 0: **** en C..F
    g[0][C]=g[0][D]=g[0][E]=g[0][F]="*"
    # Bajada por columnas C,D,E,F
    g[1][C]="*"
    g[2][D]="*"
    g[3][E]="*"
    g[4][F]="*"
    # Final: "*  *" (C y F) y "* *" (C y E)
    g[5][C]=g[5][F]="*"
    g[6][C]=g[6][E]="*"
    return g

def flatten(figs_dict):
    rows=[]
    for fig in sorted(figs_dict):
        g=figs_dict[fig]
        g=[(row+[""]*6)[:6] for row in g]
        rows.append([fig]+g[0])
        for r in g[1:]:
            rows.append([""]+r)
    return rows

def write_csv(path, rows):
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(COLS)
        w.writerows(rows)

def write_xlsx(path, rows):
    if pd is None:
        print(f"⚠️ No se creó '{path.name}' (falta pandas).")
        return
    try:
        import openpyxl  # noqa
    except Exception:
        print(f"⚠️ No se creó '{path.name}' (falta openpyxl).")
        return
    df = pd.DataFrame(rows, columns=COLS)
    with pd.ExcelWriter(path, engine="openpyxl") as w:
        sheet = path.stem
        df.to_excel(w, index=False, sheet_name=sheet)
        ws = w.sheets[sheet]
        # Estética compacta
        try:
            from openpyxl.utils import get_column_letter
            from openpyxl.styles import Font, Alignment, Border, Side
            for i,c in enumerate(COLS, start=1):
                width = 6 if c=="figura" else 5
                ws.column_dimensions[get_column_letter(i)].width = width
            font9 = Font(name="Calibri", sz=9)
            center = Alignment(horizontal="center", vertical="center")
            thin = Side(style="thin", color="DDDDDD")
            border = Border(left=thin,right=thin,top=thin,bottom=thin)
            for r in range(2, ws.max_row+1):
                ws.row_dimensions[r].height = 13.5
                for c in range(1,8):
                    cell = ws.cell(r,c)
                    cell.font = font9
                    cell.alignment = center
                    if c>=2: cell.border = border
        except Exception:
            pass

def main():
    figs = read_existing()
    # Completa 1..100 con el patrón; si ya existe una figura, la respeta
    for fig in range(1, TARGET+1):
        if fig not in figs:
            figs[fig] = pattern_grid()
    rows = flatten(figs)
    write_csv(OUT_CSV, rows)
    write_xlsx(OUT_XLSX, rows)
    print("\n✅ Listo. 100 figuras con el patrón fijado.")
    print("   Archivos: todo_fig.csv, todo_fig.xlsx")
    print("\nVer en consola (modo 'tiles'):")
    print("   python leer_figuras.py --input todo_fig.xlsx --render tiles --no-trim-h")

if __name__ == "__main__":
    main()
