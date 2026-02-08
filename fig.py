#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
fig.py — Genera figuras 39–100 con 4 '*' aleatorios por figura, sin repetir patrones.
Solo usa openpyxl (sin pandas).

Formato esperado (hoja activa del Excel):
- Fila 1: encabezados, primera columna = número de figura (puede llamarse 'fig' u otro), luego columnas b..g
- Cada figura ocupa 4 filas; en la primera fila del bloque suele venir el número y en las siguientes puede venir vacío.

Salida: escribe un nuevo archivo con las 100 figuras.
"""

import sys
import random
from pathlib import Path

try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment, Font, Border, Side
except ImportError:
    print("❌ Falta 'openpyxl'. Instala con:  pip install openpyxl")
    sys.exit(1)

# ----------------- Parámetros editables -----------------
INPUT_XLSX  = "todo_fig.xlsx"              # archivo de entrada
OUTPUT_XLSX = "todo_fig_100.xlsx"          # archivo de salida
KEEP_UPTO   = 38                           # mantener intactas 1..KEEP_UPTO
TOTAL_FIGS  = 100                          # generar hasta esta figura
ROWS_PER_FIG = 4                           # alto de cada figura (filas)
COL_LABELS   = ["b","c","d","e","f","g"]   # columnas de la cuadrícula
# --------------------------------------------------------


def read_sheet_to_rows(path: Path):
    """Devuelve (headers:list[str], rows:list[list[str]]) de la hoja activa."""
    if not Path(path).exists():
        print(f"❌ No existe el archivo: {path}")
        sys.exit(1)
    wb = load_workbook(path)
    ws = wb.active
    # encabezado
    headers = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        headers.append(str(v).strip() if v is not None else "")
    # filas
    rows = []
    for r in range(2, ws.max_row + 1):
        row = []
        all_none = True
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            s = "" if v is None else str(v).strip()
            if s != "":
                all_none = False
            row.append(s)
        if not all_none:
            rows.append(row)
    return headers, rows


def group_figs(headers, rows):
    """
    Agrupa filas en dict fig -> bloque (lista de filas con solo columnas b..g).
    Detecta:
      - primera columna = número de figura (texto/num). Las filas con esa col vacía
        se consideran continuación del mismo bloque.
      - Si la primera columna repite el número en cada fila, igual agrupa por cambios.
    """
    if not headers:
        print("❌ El archivo no tiene encabezados en la fila 1.")
        sys.exit(1)

    # Índices de columnas: primera = figura; resto buscamos b..g por nombre
    idx_fig = 0
    # Intentar encontrar b..g por nombre; si no, asumir desde la 2 hasta la 7
    name_to_idx = {headers[i].lower(): i for i in range(len(headers))}
    cols_idx = []
    for label in COL_LABELS:
        if label.lower() in name_to_idx:
            cols_idx.append(name_to_idx[label.lower()])
    if len(cols_idx) != len(COL_LABELS):
        # fallback: usar columnas 2..7 (1-based) si existen
        cols_idx = [i for i in range(1, min(7, len(headers)))]

    figs = {}
    current_fig = None
    for row in rows:
        # Asegurar ancho
        row += [""] * (max([idx_fig] + cols_idx) + 1 - len(row))
        head = row[idx_fig].strip() if idx_fig < len(row) and row[idx_fig] is not None else ""
        # si es un número, comienza nueva figura
        if head.isdigit():
            current_fig = int(head)
            figs.setdefault(current_fig, [])
            # guardar solo b..g
            figs[current_fig].append([row[i] if i < len(row) else "" for i in cols_idx])
        else:
            # continuación del bloque actual
            if current_fig is None:
                # si viene una fila previa sin fig, ignórala
                continue
            figs[current_fig].append([row[i] if i < len(row) else "" for i in cols_idx])

    # recorta colas vacías en cada bloque y asegura alto
    clean = {}
    for k, block in figs.items():
        blk = [r[:] for r in block]
        # recorta filas vacías al final
        while blk and all(c == "" for c in blk[-1]):
            blk.pop()
        if not blk:
            continue
        # si el bloque no es múltiplo de ROWS_PER_FIG o es mayor, intentamos tomar las primeras ROWS_PER_FIG filas
        if len(blk) >= ROWS_PER_FIG:
            blk = blk[:ROWS_PER_FIG]
        else:
            # completar con filas vacías para llegar a ROWS_PER_FIG
            for _ in range(ROWS_PER_FIG - len(blk)):
                blk.append([""] * len(cols_idx))
        clean[k] = blk
    return clean, cols_idx


def block_to_pattern(block):
    """Convierte un bloque (lista de filas b..g) a un patrón hashable (tuple(tuple(...)))."""
    return tuple(tuple(cell for cell in row) for row in block)


def random_block(rows=ROWS_PER_FIG, cols=len(COL_LABELS), stars=4):
    """Genera un bloque vacío y coloca 'stars' '*' en posiciones aleatorias no repetidas."""
    if stars > rows * cols:
        raise ValueError("Demasiadas '*' para el tamaño del bloque.")
    blk = [[""] * cols for _ in range(rows)]
    positions = random.sample(range(rows * cols), stars)
    for p in positions:
        r = p // cols
        c = p % cols
        blk[r][c] = "*"
    return blk


def write_output(path, headers, figs_dict, cols_idx):
    """Escribe en Excel (openpyxl) con encabezados y bloques. Figura en 1ª col, b..g después."""
    wb = Workbook()
    ws = wb.active
    ws.title = Path(path).stem

    # Encabezados: primera col = 'fig' (si ya venía, respetar nombre), luego nombres de columnas b..g
    first_col_name = headers[0] if headers and headers[0] else "fig"
    out_headers = [first_col_name] + COL_LABELS[:len(cols_idx)]
    for j, h in enumerate(out_headers, start=1):
        ws.cell(1, j).value = h

    # Estética liviana
    font = Font(name="Calibri", size=9)
    center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Escribir bloques en orden
    r = 2
    for fig in sorted(figs_dict):
        blk = figs_dict[fig]
        # asegurar alto ROWS_PER_FIG y ancho
        for i in range(ROWS_PER_FIG):
            row_vals = blk[i] if i < len(blk) else [""] * len(cols_idx)
            # Número de figura solo en la primera fila del bloque
            ws.cell(r + i, 1).value = fig if i == 0 else ""
            # celdas b..g
            for j, val in enumerate(row_vals, start=2):
                ws.cell(r + i, j).value = val
                ws.cell(r + i, j).font = font
                ws.cell(r + i, j).alignment = center
                ws.cell(r + i, j).border = border
            # aplicar estilo también a la 1ª col
            ws.cell(r + i, 1).font = font
            ws.cell(r + i, 1).alignment = center
        r += ROWS_PER_FIG

    # Anchos de columna
    for j in range(1, len(out_headers) + 1):
        ws.column_dimensions[get_column_letter(j)].width = 6 if j == 1 else 5
    # Altura de filas compacta
    for i in range(2, r):
        ws.row_dimensions[i].height = 14

    wb.save(path)


def main():
    random.seed()  # aleatoriedad real
    headers, rows = read_sheet_to_rows(INPUT_XLSX)
    figs, cols_idx = group_figs(headers, rows)

    if not figs:
        print("❌ No se detectaron figuras válidas en el archivo de entrada.")
        sys.exit(1)

    # 1) Mantener 1..KEEP_UPTO tal cual
    result = {}
    for f in sorted(figs):
        if f <= KEEP_UPTO:
            result[f] = figs[f][:]  # copia superficial

    # 2) Registrar patrones existentes para evitar repeticiones
    patterns = set()
    for f, blk in figs.items():
        pat = block_to_pattern(blk)
        patterns.add(pat)

    # 3) Generar 39..TOTAL_FIGS con 4 '*' aleatorios sin repetir
    #    (si KEEP_UPTO no es 38, genera a partir de KEEP_UPTO+1)
    start = max(KEEP_UPTO + 1, max(figs.keys()) + 1 if max(figs.keys()) > KEEP_UPTO else KEEP_UPTO + 1)
    # Pero si ya existen algunas > KEEP_UPTO, respetarlas y solo completar hasta TOTAL_FIGS
    for f in sorted(figs):
        if f > KEEP_UPTO:
            result[f] = figs[f][:]
    next_fig = max(result.keys()) + 1 if result else 1

    tries_limit = 20000  # por seguridad
    while len([k for k in result if k >= 1]) < TOTAL_FIGS:
        if tries_limit <= 0:
            print("⚠️ Se alcanzó el límite de intentos generando patrones únicos. Salida parcial.")
            break
        blk = random_block(rows=ROWS_PER_FIG, cols=len(cols_idx), stars=4)
        pat = block_to_pattern(blk)
        if pat in patterns:
            tries_limit -= 1
            continue
        patterns.add(pat)
        result[next_fig] = blk
        next_fig += 1

    # 4) Escribir salida
    write_output(OUTPUT_XLSX, headers, result, cols_idx)
    print("✅ Listo.")
    print(f"   Figuras mantenidas: 1..{KEEP_UPTO}")
    print(f"   Total figuras en salida: {len(result)}")
    print(f"   Archivo: {OUTPUT_XLSX}")

if __name__ == "__main__":
    main()
