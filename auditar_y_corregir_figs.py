#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
auditar_y_corregir_figs.py
- Lee 'todo_fig.xlsx' (solo openpyxl).
- Mantiene 1..38 tal cual (no los modifica).
- Asegura 39..100: 4 filas por figura, 6 columnas (b..g), exactamente 4 '*', sin patrones repetidos.
- Guarda 'todo_fig_100.xlsx' y 'reporte_figs.txt'.

Requisitos: pip install openpyxl
"""

import sys, random
from pathlib import Path

try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment, Font, Border, Side
except ImportError:
    print("❌ Falta 'openpyxl'. Instala con:  pip install openpyxl")
    sys.exit(1)

# Parámetros
INPUT_XLSX   = "todo_fig.xlsx"
OUTPUT_XLSX  = "todo_fig_100.xlsx"
REPORT_TXT   = "reporte_figs.txt"

KEEP_UPTO    = 38           # mantener intactas 1..38
TOTAL_FIGS   = 100          # objetivo
ROWS_PER_FIG = 4
LABELS_6     = ["b","c","d","e","f","g"]

# ---------------- utilidades ----------------
def read_sheet(path: Path):
    wb = load_workbook(path)
    ws = wb.active
    # encabezados (fila 1)
    headers = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        headers.append((str(v).strip() if v is not None else ""))
    # filas (desde fila 2)
    rows = []
    for r in range(2, ws.max_row + 1):
        row = []
        all_empty = True
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            s = "" if v is None else str(v).strip()
            if s != "":
                all_empty = False
            row.append(s)
        if not all_empty:
            rows.append(row)
    return headers, rows

def group_figs(headers, rows, log):
    """Devuelve (figs: dict[int -> block[list[list[str]]]], cols_idx:list[int], first_col_name:str)."""
    if not headers:
        raise RuntimeError("El archivo no tiene encabezados en la fila 1.")

    # Primera columna = número de figura
    idx_fig = 0
    first_col_name = headers[0] or "fig"

    # Intentar ubicar b..g por nombre (case-insensitive, espacios fuera)
    name_to_idx = {headers[i].strip().lower(): i for i in range(len(headers))}
    cols_idx = []
    for lab in LABELS_6:
        if lab.lower() in name_to_idx:
            cols_idx.append(name_to_idx[lab.lower()])

    # Fallback: si no encontramos 6, usar columnas 2..7 (1-based) que existan
    if len(cols_idx) != 6:
        cols_idx = [i for i in range(1, min(len(headers), 7))]
        log.append(f"⚠️ No se localizaron exactamente b..g por nombre; usando columnas 2..7 (encontradas {len(cols_idx)}).")

    figs = {}
    current = None
    for row in rows:
        # asegurar ancho
        need_len = max([idx_fig] + cols_idx) + 1
        if len(row) < need_len:
            row += [""] * (need_len - len(row))
        head = row[idx_fig].strip()
        if head.isdigit():
            current = int(head)
            figs.setdefault(current, [])
            figs[current].append([row[i] for i in cols_idx])
        else:
            if current is None:
                continue
            figs[current].append([row[i] for i in cols_idx])

    # limpiar colas vacías y ajustar alto
    clean = {}
    for k, blk in figs.items():
        b = [r[:] for r in blk]
        while b and all(c == "" for c in b[-1]):
            b.pop()
        if not b:
            continue
        # Si tiene más de ROWS_PER_FIG, recortamos; si menos, completamos con vacías
        if len(b) >= ROWS_PER_FIG:
            b = b[:ROWS_PER_FIG]
        else:
            b += [[""] * len(cols_idx) for _ in range(ROWS_PER_FIG - len(b))]
        clean[k] = b
    return clean, cols_idx, first_col_name

def block_pattern(blk):
    return tuple(tuple(c for c in row) for row in blk)

def normalize_block(blk):
    """Convierte cualquier no-vacío distinto de '*' en '*' (solo para comparación)."""
    out = []
    for row in blk:
        out.append([("*" if (c != "" and c != "*") else c) for c in row])
    return out

def count_stars(blk):
    return sum(1 for r in blk for c in r if c == "*")

def random_block(rows, cols, stars=4):
    blk = [[""] * cols for _ in range(rows)]
    positions = random.sample(range(rows * cols), stars)
    for p in positions:
        r = p // cols
        c = p % cols
        blk[r][c] = "*"
    return blk

def write_output(path, first_col_name, cols_idx, figs_dict):
    wb = Workbook()
    ws = wb.active
    ws.title = Path(path).stem

    # encabezados
    headers_out = [first_col_name] + [LABELS_6[i] if i < len(LABELS_6) else f"col{j}"
                                      for j,i in enumerate(range(len(cols_idx)), start=2)]
    # respeta nombre de columnas de salida como b..g
    headers_out = [first_col_name] + LABELS_6[:len(cols_idx)]
    for j, h in enumerate(headers_out, start=1):
        ws.cell(1, j).value = h

    font = Font(name="Calibri", size=9)
    center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    r = 2
    for fig in sorted(figs_dict):
        blk = figs_dict[fig]
        # 4 filas por figura
        for i in range(ROWS_PER_FIG):
            ws.cell(r + i, 1).value = fig if i == 0 else ""
            for j, val in enumerate(blk[i], start=2):
                ws.cell(r + i, j).value = val
                ws.cell(r + i, j).font = font
                ws.cell(r + i, j).alignment = center
                ws.cell(r + i, j).border = border
            ws.cell(r + i, 1).font = font
            ws.cell(r + i, 1).alignment = center
        r += ROWS_PER_FIG

    # anchos y alturas
    for j in range(1, 8):
        ws.column_dimensions[get_column_letter(j)].width = 6 if j == 1 else 5
    for i in range(2, r):
        ws.row_dimensions[i].height = 14

    wb.save(path)

# ---------------- flujo principal ----------------
def main():
    random.seed()
    path_in = Path(INPUT_XLSX)
    if not path_in.exists():
        print(f"❌ No existe '{INPUT_XLSX}'.")
        sys.exit(1)

    log = []
    headers, rows = read_sheet(path_in)
    figs_in, cols_idx, first_col = group_figs(headers, rows, log)

    if not figs_in:
        print("❌ No se detectaron figuras válidas en el archivo.")
        sys.exit(1)

    # 1) preservar 1..KEEP_UPTO tal cual
    result = {}
    for f in sorted(figs_in):
        if f <= KEEP_UPTO:
            result[f] = [r[:] for r in figs_in[f]]

    # 2) patrones ya existentes (normalizados) para evitar duplicados
    patterns = set()
    for f, blk in figs_in.items():
        pat = block_pattern(normalize_block(blk))
        patterns.add(pat)

    # 3) revisar y reportar inconsistencias 1..KEEP_UPTO (no modificar)
    for f in range(1, min(KEEP_UPTO, max(figs_in.keys())) + 1):
        if f not in figs_in:
            log.append(f"⚠️ Falta figura {f} en entrada.")
            continue
        blk = figs_in[f]
        stars = count_stars(normalize_block(blk))
        if stars != 4:
            log.append(f"ℹ️ Figura {f}: tiene {stars} '*' (no se modifica por estar en 1..{KEEP_UPTO}).")

    # 4) copiar también las ya existentes > KEEP_UPTO (si las hay), normalizándolas a 4 filas
    existing_above = [f for f in figs_in if f > KEEP_UPTO]
    for f in existing_above:
        blk = figs_in[f]
        # normalizamos a 4 filas (ya viene de group_figs)
        # si no tiene 4 '*', lo corregimos
        nb = normalize_block(blk)
        stars = count_stars(nb)
        if stars != 4:
            # reemplazar por un bloque nuevo único
            tries = 20000
            while tries > 0:
                cand = random_block(ROWS_PER_FIG, len(cols_idx), stars=4)
                pat = block_pattern(cand)
                if pat not in patterns:
                    patterns.add(pat)
                    nb = cand
                    log.append(f"✔️ Figura {f}: corregida a 4 '*' (antes {stars}).")
                    break
                tries -= 1
            if tries == 0:
                log.append(f"❗ No se pudo asignar patrón único a figura {f}. (Se dejó como estaba)")
        result[f] = nb

    # 5) completar hasta TOTAL_FIGS con bloques únicos de 4 '*'
    next_fig = max(result.keys()) + 1 if result else 1
    while len(result) < TOTAL_FIGS:
        # si hay huecos intermedios, cúbrelos con su número
        while next_fig in result:
            next_fig += 1
        tries = 20000
        while tries > 0:
            cand = random_block(ROWS_PER_FIG, len(cols_idx), stars=4)
            pat = block_pattern(cand)
            if pat not in patterns:
                patterns.add(pat)
                result[next_fig] = cand
                break
            tries -= 1
        if tries == 0:
            log.append("❗ Límite alcanzado generando patrones únicos. Salida parcial.")
            break
        next_fig += 1

    # 6) escribir salida y reporte
    write_output(OUTPUT_XLSX, first_col, cols_idx, result)
    with open(REPORT_TXT, "w", encoding="utf-8") as f:
        f.write("REPORTE DE AUDITORÍA Y CORRECCIÓN\n")
        f.write(f"Archivo base : {INPUT_XLSX}\n")
        f.write(f"Archivo salida: {OUTPUT_XLSX}\n\n")
        for line in log:
            f.write(line + "\n")
        f.write("\nResumen:\n")
        f.write(f"- Figuras preservadas 1..{KEEP_UPTO}\n")
        f.write(f"- Total figuras en salida: {len(result)}\n")

    print("✅ Listo.")
    print(f"   Guardado: {OUTPUT_XLSX}")
    print(f"   Reporte : {REPORT_TXT}")
    if log:
        print("   (Hay avisos; revisa reporte_figs.txt)")

if __name__ == "__main__":
    main()
