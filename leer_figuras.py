#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
leer_figuras.py — visor en consola (solo openpyxl)
- 1ª columna: número de figura
- Rejilla: columnas b..g (si no existen por nombre, usa columnas 2..7)
- Cada figura ocupa 4 filas

Uso:
  python leer_figuras.py --input todo_fig_100_validado.xlsx
  python leer_figuras.py --input todo_fig_100_validado.xlsx --max 20
  python leer_figuras.py --input todo_fig_100_validado.xlsx --only 9
  python leer_figuras.py --input archivo.csv   (también soporta CSV)
  python leer_figuras.py --render boxes        (cajitas 3x3)
"""

import sys, csv, argparse
from pathlib import Path

# ---------- lectura XLSX (openpyxl) ----------
def read_xlsx(path: Path):
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("❌ Falta 'openpyxl'. Instala con: pip install openpyxl")
        sys.exit(1)
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    headers = []
    for c in range(1, ws.max_column+1):
        v = ws.cell(1, c).value
        headers.append("" if v is None else str(v).strip())
    rows = []
    for r in range(2, ws.max_row+1):
        row=[]; all_empty=True
        for c in range(1, ws.max_column+1):
            v = ws.cell(r, c).value
            s = "" if v is None else str(v).strip()
            if s != "": all_empty=False
            row.append(s)
        if not all_empty: rows.append(row)
    return headers, rows

# ---------- lectura CSV ----------
def read_csv(path: Path):
    with path.open(newline="", encoding="utf-8-sig") as f:
        data = [ [ (x or "").strip() for x in row ] for row in csv.reader(f) ]
    if not data: return [], []
    headers = data[0]
    rows = [r for r in data[1:] if any(c != "" for c in r)]
    return headers, rows

# ---------- agrupar por figura ----------
LABELS = ["b","c","d","e","f","g"]
ROWS_PER_FIG = 4

def group_figs(headers, rows):
    if not headers: raise RuntimeError("No hay encabezados en la fila 1.")
    idx_fig = 0  # 1ª columna es el número
    # localizar b..g por nombre; si no, usar 2..7
    name_to_idx = {headers[i].strip().lower(): i for i in range(len(headers))}
    cols_idx=[]
    for lab in LABELS:
        if lab.lower() in name_to_idx:
            cols_idx.append(name_to_idx[lab.lower()])
    if len(cols_idx) != 6:
        cols_idx = [i for i in range(1, min(7, len(headers)))]  # fallback

    figs = {}
    cur = None
    for row in rows:
        need = max([idx_fig]+cols_idx)+1
        if len(row) < need: row += [""]*(need-len(row))
        head = row[idx_fig]
        if head.isdigit():
            cur = int(head)
            figs.setdefault(cur, [])
            figs[cur].append([row[i] for i in cols_idx])
        else:
            if cur is None: continue
            figs[cur].append([row[i] for i in cols_idx])

    # normaliza alto a 4 filas y recorta colas vacías
    clean={}
    for f, blk in figs.items():
        b=[r[:] for r in blk]
        while b and all(c=="" for c in b[-1]): b.pop()
        if not b: continue
        if len(b) >= ROWS_PER_FIG: b = b[:ROWS_PER_FIG]
        else: b += [[""]*len(cols_idx) for _ in range(ROWS_PER_FIG-len(b))]
        clean[f]=b
    return clean

# ---------- render ----------
def trim_grid(grid, trim_h=True, trim_v=True):
    if not grid: return []
    g=[r[:] for r in grid]
    if trim_v:
        top,bottom=0,len(g)
        while top<bottom and all(c=="" for c in g[top]): top+=1
        while bottom>top and all(c=="" for c in g[bottom-1]): bottom-=1
        g=g[top:bottom] if top<bottom else []
        if not g: return []
    if trim_h:
        cols=len(g[0]); left,right=0,cols
        def col_empty(ci): return all(row[ci]=="" for row in g)
        while left<right and col_empty(left): left+=1
        while right>left and col_empty(right-1): right-=1
        if right<=left: return []
        g=[row[left:right] for row in g]
    return g

def render_tiles_row(row, gap):
    sep = " "*gap
    return sep.join(c if c!="" else "" for c in row)

def render_boxes_row(row, gap):
    def box(v):
        s = v if v!="" else " "
        return ["┌─┐", f"│{s[:2]}│", "└─┘"]
    sep=" "*gap
    l1=l2=l3=""
    for i,v in enumerate(row):
        b=box(v)
        if i==0: l1,l2,l3=b
        else:
            l1=f"{l1}{sep}{b[0]}"
            l2=f"{l2}{sep}{b[1]}"
            l3=f"{l3}{sep}{b[2]}"
    return [l1,l2,l3]

def print_figure(fig, grid, mode="tiles", gap=1, trim_h=True, trim_v=True):
    g = trim_grid(grid, trim_h, trim_v)
    if not g: return
    print(f"\nFigura {fig}")
    for row in g:
        if mode=="boxes":
            a,b,c = render_boxes_row(row, gap)
            print(a); print(b); print(c)
        else:
            print(render_tiles_row(row, gap))
    print("")

# ---------- CLI ----------
def main():
    ap = argparse.ArgumentParser(description="Visor de figuras en consola (XLSX/CSV).")
    ap.add_argument("--input","-i", default="todo_fig_100_validado.xlsx")
    ap.add_argument("--render", choices=["tiles","boxes"], default="tiles")
    ap.add_argument("--gap", type=int, default=1)
    ap.add_argument("--only", type=int, help="Mostrar solo esta figura")
    ap.add_argument("--max", type=int, help="Mostrar hasta esta figura (inclusive)")
    ap.add_argument("--no-trim-h", action="store_true", help="No recortar columnas vacías")
    ap.add_argument("--no-trim-v", action="store_true", help="No recortar filas vacías")
    args = ap.parse_args()

    path = Path(args.input)
    if not path.exists():
        print(f"❌ No existe: {path}")
        sys.exit(1)

    # leer
    if path.suffix.lower()==".xlsx":
        headers, rows = read_xlsx(path)
    else:
        headers, rows = read_csv(path)

    try:
        figs = group_figs(headers, rows)
    except Exception as e:
        print(f"❌ {e}"); sys.exit(1)

    keys = sorted(figs.keys())
    if args.only is not None:
        keys = [k for k in keys if k == args.only]
    elif args.max is not None:
        keys = [k for k in keys if k <= args.max]

    if not keys:
        print("⚠️ No hay figuras para mostrar con esos filtros.")
        return

    for k in keys:
        print_figure(
            k, figs[k],
            mode=args.render,
            gap=args.gap,
            trim_h=not args.no_trim_h,
            trim_v=not args.no_trim_v
        )

if __name__ == "__main__":
    main()
