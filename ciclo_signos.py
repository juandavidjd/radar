#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ciclo_signos.py — versión plantilla-aware
- Lee 'todo_pos.csv' (numeros,signos) → genera 100 combinaciones únicas (4 dígitos).
- Asigna 1–3 signos por combinación (dígitos únicos en orden).
- Lee 'plantilla_fig.xlsx' o 'plantilla_fig.csv' y APRENDE el diseño:
    * Detecta coordenadas de dígitos (top-4 más frecuentes donde hay dígitos o '*')
    * Detecta celdas decorativas (no vacías no-dígito fuera de dígitos)
    * Calcula altura necesaria del bloque
  Si no se puede inferir, usa respaldo 3×6 con dígitos en C–D (2×2).
- Genera 100 figuras siguiendo el MISMO estilo.
- Exporta: todo_sum.*, todo_fig.*, todo_map.* (CSV + XLSX si hay pandas+openpyxl).
"""

import csv, sys, random, re
from collections import Counter, defaultdict
from pathlib import Path

# XLSX opcional
try:
    import pandas as pd
except Exception:
    pd = None

POS_PATH   = Path("todo_pos.csv")
TPL_XLSX   = Path("plantilla_fig.xlsx")
TPL_CSV    = Path("plantilla_fig.csv")

SUM_CSV    = Path("todo_sum.csv")
FIG_CSV    = Path("todo_fig.csv")
FIG_XLSX   = Path("todo_fig.xlsx")
MAP_CSV    = Path("todo_map.csv")
MAP_XLSX   = Path("todo_map.xlsx")

FIG_COLS   = ["figura","a","b","c","d","e","f"]
TARGET_FIGS= 100
COMBOS_N   = 100

DIGIT_CHAR_RE = re.compile(r"^\d$")   # exactamente 1 dígito
HAS_DIGIT_RE  = re.compile(r"\d")     # contiene algún dígito

# ------------- util CSV/XLSX -------------
def read_csv(path: Path):
    if not path.exists(): return []
    with path.open(newline="", encoding="utf-8-sig") as f:
        return [[c.strip() for c in r] for r in csv.reader(f)]

def write_csv_rows(path: Path, header, rows):
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        w=csv.writer(f); w.writerow(header); w.writerows(rows)

def write_csv_dicts(path: Path, header, dict_rows):
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        w=csv.DictWriter(f, fieldnames=header); w.writeheader(); w.writerows(dict_rows)

def write_xlsx_rows(path: Path, header, rows, col_widths=None):
    if pd is None:
        print(f"⚠️ No se creó '{path.name}' (falta pandas)."); return None
    try:
        import openpyxl  # noqa
    except Exception:
        print(f"⚠️ No se creó '{path.name}' (falta openpyxl)."); return None
    df = pd.DataFrame(rows, columns=header)
    with pd.ExcelWriter(path, engine="openpyxl") as w:
        sheet = path.stem
        df.to_excel(w, index=False, sheet_name=sheet)
        ws = w.sheets[sheet]
        # Anchos de columnas
        try:
            from openpyxl.utils import get_column_letter
            for i,c in enumerate(header, start=1):
                width = (col_widths or {}).get(c, None)
                if width is None:
                    width = max(8, min(20, max(len(c), *(len(str(x)) for x in df[c].head(400)))+2))
                ws.column_dimensions[get_column_letter(i)].width = width
        except Exception:
            pass
        return ws

# ------------- POS: numeros ↔ signos -------------
def load_num_sign_map(pos_path: Path):
    rows = read_csv(pos_path)
    if not rows:
        raise FileNotFoundError(f"No se encontró '{pos_path}'. Debe tener columnas: numeros, signos.")
    head = [h.lower() for h in rows[0]]
    try:
        i_num = head.index("numeros")
        i_sig = head.index("signos")
        data = rows[1:]
    except ValueError:
        i_num, i_sig = 0, 1
        data = rows
    mapping={}
    for r in data:
        if len(r)<=max(i_num,i_sig): continue
        d,s = r[i_num].strip(), r[i_sig].strip()
        if len(d)==1 and d.isdigit():
            mapping[int(d)] = s
    if not mapping:
        raise ValueError("No se pudo construir el mapa 0..9 -> signo desde todo_pos.csv.")
    return mapping, sorted(mapping.keys())

# ------------- 100 combinaciones -------------
def generate_4digit_combos(digits, n=100):
    combos=set(); tries=0
    while len(combos)<n and tries<n*300:
        combos.add("".join(str(random.choice(digits)) for _ in range(4)))
        tries+=1
    combos=list(combos); random.shuffle(combos)
    if len(combos)<n:
        print(f"⚠️ Solo {len(combos)} combinaciones únicas con dígitos {digits}.")
    return combos[:n]

def signs_for_combo(combo, num2sign):
    seen=[]; used=set()
    for ch in combo:
        d=int(ch)
        if d in used: continue
        used.add(d)
        s=num2sign.get(d,"")
        if s: seen.append(s)
        if len(seen)==3: break
    return seen  # 1..3

# ------------- leer plantilla & aprender estilo -------------
def read_template_grids():
    # Prefiere XLSX
    rows=None
    if TPL_XLSX.exists() and pd is not None:
        try:
            df = pd.read_excel(TPL_XLSX, dtype=str).fillna("")
            cols=list(df.columns)
            if cols: cols[0]="figura"; df.columns=cols
            for c in FIG_COLS:
                if c not in df.columns: df[c]=""
            df=df[FIG_COLS]
            rows=[[str(v) for v in r] for _,r in df.iterrows()]
        except Exception:
            rows=None
    if rows is None and TPL_CSV.exists():
        rows=read_csv(TPL_CSV)
    if not rows: return {}
    # agrupar por figura
    grids={}; fig=None
    for r in rows:
        r=(r+[""]*7)[:7]
        head=(r[0] or "").strip()
        a_f=[(r[i] or "").strip() for i in range(1,7)]
        if head.isdigit():
            fig=int(head); grids.setdefault(fig,[]); grids[fig].append(a_f)
        else:
            if fig is None: continue
            grids[fig].append(a_f)
    return grids

def learn_style_from_template(grids: dict):
    """
    Devuelve:
      rows_needed: int
      digit_coords: list[(ri,ci)] 4 coords
      decor_coords: set[(ri,ci)]  decor
    Si no se puede, retorna respaldo (3 filas, dígitos en (1,2)(1,3)(2,2)(2,3)).
    """
    # respaldo
    fallback = {
        "rows_needed": 3,
        "digit_coords": [(1,2),(1,3),(2,2),(2,3)],
        "decor_coords": {(0,1),(0,4),(2,1),(2,4),(1,0),(1,5)},
    }
    if not grids: return fallback

    # Normalizar: limitar a 6 columnas y anotar celdas
    digit_counter = Counter()
    decor_counter = Counter()
    max_row = 0

    for _fig, g in grids.items():
        # filtrar líneas vacías al final
        gg=[row[:6] + [""]*(6-len(row)) if len(row)<6 else row[:6] for row in g]
        while gg and all(x=="" for x in gg[-1]): gg.pop()
        if not gg: continue
        max_row = max(max_row, len(gg))
        # clasificar celdas
        for ri,row in enumerate(gg):
            for ci,val in enumerate(row):
                s=(val or "").strip()
                if s=="":
                    continue
                if len(s)==1 and s.isdigit():
                    digit_counter[(ri,ci)] += 1
                elif s=="*":
                    # tratar * como posible dígito placeholder
                    digit_counter[(ri,ci)] += 1
                else:
                    decor_counter[(ri,ci)] += 1

    if not digit_counter:
        return fallback

    # Top-4 coords más frecuentes para dígitos
    digit_coords = [p for p,_ in digit_counter.most_common(4)]
    if len(digit_coords)<4:
        # completar cuadrícula compacta alrededor de las más frecuentes
        base = digit_coords[0] if digit_coords else (1,2)
        r,c = base
        need=set([(r,c),(r,c+1),(r+1,c),(r+1,c+1)])
        for p in digit_coords: 
            need.discard(p)
        digit_coords = digit_coords + list(need)
        digit_coords = digit_coords[:4]

    # decor = posiciones decor frecuentes que no son dígito
    decor_coords = {p for p,_ in decor_counter.items() if p not in set(digit_coords)}

    rows_needed = max(max_row, max(r for r,c in digit_coords)+1)
    rows_needed = max(rows_needed, 3)  # al menos 3

    return {
        "rows_needed": rows_needed,
        "digit_coords": digit_coords,
        "decor_coords": decor_coords,
    }

# ------------- construir 100 figuras con el estilo aprendido -------------
def build_grid(rows_needed, digit_coords, decor_coords, combo):
    # grid vacío
    g=[[""]*6 for _ in range(rows_needed)]
    # decor
    for (ri,ci) in decor_coords:
        if 0<=ri<rows_needed and 0<=ci<6:
            g[ri][ci] = "·"  # conservador, un puntito minimal
    # dígitos
    digs=list(combo)
    # ordenar coords por fila/col para consistencia visual
    dcoords=sorted(digit_coords)
    # si hay menos de 4 coords, duplicamos las últimas (no debería pasar)
    while len(dcoords)<4: dcoords.append(dcoords[-1])
    for k,(ri,ci) in enumerate(dcoords[:4]):
        if 0<=ri<rows_needed and 0<=ci<6:
            g[ri][ci] = digs[k]
    return g

# ------------- MAIN -------------
def main():
    # 1) mapa números-signos
    try:
        num2sign, digits = load_num_sign_map(POS_PATH)
    except Exception as e:
        print(f"❌ {e}"); sys.exit(1)

    # 2) 100 combinaciones
    combos = generate_4digit_combos(digits, n=COMBOS_N)

    # 3) signos por combinación (1–3)
    sum_rows=[]
    for c in combos:
        sigs=signs_for_combo(c, num2sign)
        sum_rows.append({
            "combinaciones": c,
            "signo 1": sigs[0] if len(sigs)>=1 else "",
            "signo 2": sigs[1] if len(sigs)>=2 else "",
            "signo 3": sigs[2] if len(sigs)>=3 else "",
        })
    write_csv_dicts(SUM_CSV, ["combinaciones","signo 1","signo 2","signo 3"], sum_rows)
    _ = write_xlsx_rows(SUM_CSV.with_suffix(".xlsx"),
                        ["combinaciones","signo 1","signo 2","signo 3"],
                        [[r["combinaciones"],r["signo 1"],r["signo 2"],r["signo 3"]] for r in sum_rows],
                        col_widths={"combinaciones":12,"signo 1":16,"signo 2":16,"signo 3":16})

    # 4) aprender estilo desde plantilla
    tpl_grids = read_template_grids()
    style = learn_style_from_template(tpl_grids)
    rows_needed = style["rows_needed"]
    digit_coords = style["digit_coords"]
    decor_coords = style["decor_coords"]

    # 5) construir 100 figuras (fig ↔ combo aleatorio)
    figs = list(range(1, TARGET_FIGS+1))
    random.shuffle(figs)
    combos2 = combos[:]
    while len(combos2)<TARGET_FIGS:
        combos2.append(random.choice(combos))
    filled={}; fmap={}
    for fig, combo in zip(figs, combos2[:TARGET_FIGS]):
        g = build_grid(rows_needed, digit_coords, decor_coords, combo)
        filled[fig]=g; fmap[fig]=combo

    # 6) guardar todo_fig.*
    rows=[]
    for fig in sorted(filled):
        g=filled[fig]
        rows.append([fig]+g[0])
        for r in g[1:]:
            rows.append([""]+r)
    write_csv_rows(FIG_CSV, FIG_COLS, rows)

    ws = write_xlsx_rows(FIG_XLSX, FIG_COLS, rows,
                         col_widths={"figura":6,"a":4.5,"b":4.5,"c":4.5,"d":4.5,"e":4.5,"f":4.5})
    if ws is not None:
        try:
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
            font9 = Font(name="Calibri", sz=9)
            center = Alignment(horizontal="center", vertical="center")
            thin = Side(style="thin", color="DDDDDD")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            # altura baja
            for r in range(2, ws.max_row+1):
                ws.row_dimensions[r].height = 13.5
            # estilo celdas
            for r in range(2, ws.max_row+1):
                for c in range(1, 8):
                    cell = ws.cell(r,c)
                    cell.font = font9
                    cell.alignment = center
                    if c>=2: cell.border = border
            # color suave a decor y distinto a dígitos
            fillDecor = PatternFill("solid", fgColor="F2F2F2")
            fillDigit = PatternFill("solid", fgColor="E7F4FF")
            # repintar por bloques de figura
            r=2
            while r<=ws.max_row:
                # determinar alto del bloque por filas sucesivas sin nueva 'figura'
                start=r
                # primera fila ya es de figura (tiene número en col A si es inicio)
                # buscar hasta la siguiente fila que tenga valor en col A (nueva figura) o fin
                r2=start
                while r2<=ws.max_row and (ws.cell(r2,1).value in (None,"")):
                    r2+=1
                # bloque actual es [start-1 .. r2-1], pero start-1 tiene número de figura
                blk_start = start-1
                blk_end   = r2-1
                # aplicar colores en celdas decor/dígito según coords relativas
                # Revisar filas del bloque: fila con número = blk_start
                for (ri,ci) in decor_coords:
                    rr = blk_start + 1 + ri  # +1 porque tras número viene primera fila de grid
                    cc = 1 + ci
                    if 2 <= rr <= ws.max_row and 2 <= cc <= 7:
                        ws.cell(rr,cc).fill = fillDecor
                for (ri,ci) in digit_coords:
                    rr = blk_start + 1 + ri
                    cc = 1 + ci
                    if 2 <= rr <= ws.max_row and 2 <= cc <= 7:
                        ws.cell(rr,cc).fill = fillDigit
                r = r2 if r2>start else r2+1
        except Exception:
            pass

    # 7) mapa figura↔combinación
    map_rows = [[k, v] for k,v in sorted(fmap.items())]
    write_csv_rows(MAP_CSV, ["figura","combinacion"], map_rows)
    write_xlsx_rows(MAP_XLSX, ["figura","combinacion"], map_rows,
                    col_widths={"figura":8,"combinacion":12})

    # Resumen
    print("\n✅ Listo.")
    print(f"   Dígitos disponibles: {sorted(digits)}")
    print(f"   Generadas {len(combos)} combinaciones únicas.")
    print(f"   Estilo aprendido: {rows_needed} filas; dígitos en {sorted(digit_coords)}; decor={len(decor_coords)} celdas.")
    print("   Archivos:")
    print("   - todo_sum.csv / .xlsx")
    print("   - todo_fig.csv / .xlsx  (100 figuras con el estilo de la plantilla)")
    print("   - todo_map.csv / .xlsx")
    print("\nVer en consola:")
    print("  python leer_figuras.py --input todo_fig.xlsx  (ajusta ALLOWED_MAX_FIG=100 en tu lector)")

if __name__ == "__main__":
    main()
