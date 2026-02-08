#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
verificar_figs.py
Valida figuras en XLSX agrupando correctamente 4 filas por figura y rejilla b..g.

- Cuenta que cada figura tenga EXACTAMENTE 4 '*'
- Detecta patrones duplicados (mismo 4x6 de '*' y vacíos)
- Muestra resumen y, si hay problemas, lista algunos ejemplos

Uso:
  python verificar_figs.py --input todo_fig_100_validado.xlsx
Requisitos: pip install openpyxl
"""

import sys, argparse
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("❌ Falta 'openpyxl'. Instala con: pip install openpyxl")
    sys.exit(1)

ROWS_PER_FIG = 4
GRID_COLS_EXPECTED = ["b","c","d","e","f","g"]

def read_sheet(path: Path):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    # encabezados
    headers = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        headers.append("" if v is None else str(v).strip())
    # filas
    rows = []
    for r in range(2, ws.max_row + 1):
        row=[]
        all_empty=True
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            s = "" if v is None else str(v).strip()
            if s != "": all_empty = False
            row.append(s)
        if not all_empty:
            rows.append(row)
    return headers, rows

def group_figs(headers, rows, log):
    if not headers:
        raise RuntimeError("No hay encabezados en la fila 1.")
    idx_fig = 0
    # localizar columnas b..g por nombre; si no, usar 2..7
    name_to_idx = {headers[i].strip().lower(): i for i in range(len(headers))}
    cols_idx=[]
    for lab in GRID_COLS_EXPECTED:
        if lab.lower() in name_to_idx:
            cols_idx.append(name_to_idx[lab.lower()])
    if len(cols_idx) != 6:
        cols_idx = [i for i in range(1, min(7, len(headers)))]
        log.append("⚠️ No hallé exactamente b..g; uso columnas 2..7 como rejilla.")

    figs={}
    current=None
    for row in rows:
        need = max([idx_fig]+cols_idx)+1
        if len(row) < need:
            row += [""]*(need-len(row))
        head = row[idx_fig]
        if head.isdigit():
            current = int(head)
            figs.setdefault(current, [])
            figs[current].append([row[i] for i in cols_idx])
        else:
            if current is None:
                continue
            figs[current].append([row[i] for i in cols_idx])

    # normalizar a 4 filas
    clean={}
    for f, blk in figs.items():
        b = [r[:] for r in blk]
        while b and all(c=="" for c in b[-1]):
            b.pop()
        if not b: continue
        if len(b) >= ROWS_PER_FIG:
            b = b[:ROWS_PER_FIG]
        else:
            b += [[""]*len(cols_idx) for _ in range(ROWS_PER_FIG - len(b))]
        clean[f] = b
    return clean, cols_idx

def norm_block(block):
    # cualquier celda no vacía que no sea '*' se considera '*'
    return [[("*" if (c != "" and c != "*") else c) for c in row] for row in block]

def block_key(block):
    return tuple(tuple(row) for row in block)

def count_stars(block):
    return sum(1 for r in block for c in r if c == "*")

def main():
    ap = argparse.ArgumentParser(description="Validador de figuras (agrupa 4 filas por figura).")
    ap.add_argument("--input", "-i", default="todo_fig_100_validado.xlsx")
    args = ap.parse_args()

    path = Path(args.input)
    if not path.exists():
        print(f"❌ No existe: {path}")
        sys.exit(1)

    log=[]
    headers, rows = read_sheet(path)
    figs, cols_idx = group_figs(headers, rows, log)
    if not figs:
        print("⚠️ No se detectaron figuras."); sys.exit(0)

    # validación
    patterns = {}
    pat_to_figs = {}
    bad_star = []
    for f, blk in sorted(figs.items()):
        nb = norm_block(blk)
        stars = count_stars(nb)
        if stars != 4:
            bad_star.append((f, stars))
        key = block_key(nb)
        patterns[f] = key
        pat_to_figs.setdefault(key, []).append(f)

    dups = {k:v for k,v in pat_to_figs.items() if len(v) > 1}

    # resumen
    print(f"Archivo: {path.name}")
    print(f"Figuras detectadas: {len(figs)}")
    print(f"Rejilla columnas: {len(cols_idx)} (esperado 6: b..g)")
    if log:
        for line in log: print(line)

    if not bad_star:
        print("✔ Todas las figuras tienen exactamente 4 '*'.")
    else:
        print(f"❗ Figuras con conteo distinto de 4 '*': {len(bad_star)}")
        print("   Ejemplos:", bad_star[:10])

    if not dups:
        print("✔ No hay patrones duplicados.")
    else:
        print(f"❗ Patrones duplicados: {len(dups)} grupo(s). Ejemplos:")
        shown=0
        for k, figs_list in dups.items():
            print("   →", sorted(figs_list))
            shown += 1
            if shown >= 8: break

if __name__ == "__main__":
    main()
