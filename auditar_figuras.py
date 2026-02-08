#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
auditar_figuras.py
Revisa cada figura en 'todo_fig_100.xlsx' (sin pandas):
- 1ª columna = número de figura (texto/num)
- Columnas b..g = rejilla
- Cada figura ocupa 4 filas (bloque)
Checks:
  • 4 filas por figura
  • exactamente 4 '*'
  • patrones únicos (no repetidos)
Opcional --fix:
  • corrige SOLO figuras > 38 (1..38 nunca se tocan)
  • asigna patrón aleatorio válido (4 '*') no repetido

Uso:
  python auditar_figuras.py                  # solo audita, no cambia
  python auditar_figuras.py --fix            # audita y corrige >38
Requiere: pip install openpyxl
"""

import sys, random, argparse
from pathlib import Path

try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment, Font, Border, Side
except ImportError:
    print("❌ Falta 'openpyxl'. Instala con:  pip install openpyxl")
    sys.exit(1)

INPUT_XLSX   = "todo_fig_100.xlsx"
OUTPUT_XLSX  = "todo_fig_100_validado.xlsx"
REPORT_TXT   = "reporte_figs.txt"

ROWS_PER_FIG = 4
LABELS_6     = ["b","c","d","e","f","g"]
LOCK_MAX_INT = 38        # hasta aquí NO se corrige (solo reporte)
TOTAL_FIGS_EXPECTED = 100

# ------------- utilidades -------------
def read_sheet(path: Path):
    wb = load_workbook(path)
    ws = wb.active
    headers = []
    for c in range(1, ws.max_column+1):
        v = ws.cell(1, c).value
        headers.append("" if v is None else str(v).strip())
    rows = []
    for r in range(2, ws.max_row+1):
        row=[]; all_empty=True
        for c in range(1, ws.max_column+1):
            v = ws.cell(r, c).value
            s = "" if v is None else str(v).strip()
            if s != "": all_empty=False
            row.append(s)
        if not all_empty:
            rows.append(row)
    return headers, rows

def group_figs(headers, rows, log):
    # 1ª columna = figura; resto la rejilla
    if not headers: raise RuntimeError("No hay encabezados en la fila 1.")
    idx_fig = 0

    # localizar b..g por nombre; fallback 2..7
    name_to_idx = {headers[i].strip().lower(): i for i in range(len(headers))}
    cols_idx=[]
    for lab in LABELS_6:
        if lab.lower() in name_to_idx:
            cols_idx.append(name_to_idx[lab.lower()])
    if len(cols_idx) != 6:
        cols_idx = [i for i in range(1, min(7, len(headers)))]
        log.append("⚠️ No hallé todas b..g; uso columnas 2..7 como rejilla.")

    figs = {}
    cur=None
    for row in rows:
        need = max([idx_fig]+cols_idx)+1
        if len(row)<need: row += [""]*(need-len(row))
        head = row[idx_fig]
        if head.isdigit():
            cur = int(head)
            figs.setdefault(cur, [])
            figs[cur].append([row[i] for i in cols_idx])
        else:
            if cur is None: continue
            figs[cur].append([row[i] for i in cols_idx])

    # normaliza alto a 4 filas
    clean={}
    for k,blk in figs.items():
        b=[r[:] for r in blk]
        # recorta colas vacías
        while b and all(c=="" for c in b[-1]): b.pop()
        if not b: continue
        if len(b)>=ROWS_PER_FIG: b=b[:ROWS_PER_FIG]
        else:
            b += [[""]*len(cols_idx) for _ in range(ROWS_PER_FIG-len(b))]
        clean[k]=b
    return clean, cols_idx, headers[0] or "fig"

def normalize_block(blk):
    return [[("*" if (c!="" and c!="*") else c) for c in row] for row in blk]

def pattern_tuple(blk):
    return tuple(tuple(c for c in row) for row in blk)

def count_stars(blk):
    return sum(1 for r in blk for c in r if c=="*")

def random_block(rows, cols, stars=4):
    blk=[[""]*cols for _ in range(rows)]
    pos = random.sample(range(rows*cols), stars)
    for p in pos:
        r=p//cols; c=p%cols
        blk[r][c]="*"
    return blk

def write_output(path, first_col_name, figs_dict, cols_count):
    wb=Workbook(); ws=wb.active; ws.title=Path(path).stem
    headers_out=[first_col_name]+LABELS_6[:cols_count]
    for j,h in enumerate(headers_out, start=1): ws.cell(1,j).value=h

    font=Font(name="Calibri", size=9)
    center=Alignment(horizontal="center", vertical="center")
    thin=Side(style="thin", color="DDDDDD")
    border=Border(left=thin,right=thin,top=thin,bottom=thin)

    r=2
    for fig in sorted(figs_dict):
        blk=figs_dict[fig]
        for i in range(ROWS_PER_FIG):
            ws.cell(r+i,1).value = fig if i==0 else ""
            for j,val in enumerate(blk[i], start=2):
                cell=ws.cell(r+i, j)
                cell.value=val
                cell.font=font; cell.alignment=center; cell.border=border
            ws.cell(r+i,1).font=font; ws.cell(r+i,1).alignment=center
        r+=ROWS_PER_FIG

    # estética compacta
    from openpyxl.utils import get_column_letter
    for j in range(1, 8):
        ws.column_dimensions[get_column_letter(j)].width = 6 if j==1 else 5
    for i in range(2,r):
        ws.row_dimensions[i].height = 14
    wb.save(path)

# ------------- main -------------
def main():
    parser=argparse.ArgumentParser(description="Audita (y opcionalmente corrige) todo_fig_100.xlsx")
    parser.add_argument("--input", default=INPUT_XLSX)
    parser.add_argument("--output", default=OUTPUT_XLSX)
    parser.add_argument("--report", default=REPORT_TXT)
    parser.add_argument("--fix", action="store_true", help="Corrige figuras > 38 a 4 '*' únicos")
    args=parser.parse_args()

    random.seed()
    log=[]
    headers, rows = read_sheet(Path(args.input))
    figs, cols_idx, first_col = group_figs(headers, rows, log)
    if not figs:
        print("❌ No se detectaron figuras."); sys.exit(1)

    # Auditoría
    issues=0
    patterns=set()
    by_pat={}  # patrón -> [figs]
    for f,blk in figs.items():
        nb = normalize_block(blk)
        pat = pattern_tuple(nb)
        stars = count_stars(nb)
        by_pat.setdefault(pat, []).append(f)
        if stars != 4:
            issues += 1
            log.append(f"• Figura {f}: tiene {stars} '*' (esperado 4).")
        if len(blk) != ROWS_PER_FIG:
            issues += 1
            log.append(f"• Figura {f}: alto {len(blk)} filas (esperado {ROWS_PER_FIG}).")

    # detectar duplicados
    dups = {pat: figs_list for pat, figs_list in by_pat.items() if len(figs_list) > 1}
    for pat, fs in dups.items():
        fs_sorted=sorted(fs)
        log.append(f"• Patrón repetido en figuras: {', '.join(map(str,fs_sorted))}")

    # Correcciones (solo > 38)
    result = {f:[r[:] for r in figs[f]] for f in figs}  # copia
    if args.fix:
        # congelar patrones válidos (normalizados) que queremos reservar
        lock_patterns=set()
        for f in sorted(figs):
            nb=normalize_block(figs[f])
            lock_patterns.add(pattern_tuple(nb))

        # corrige > 38 con 4 '*' y patrón único
        cols_count = len(cols_idx)
        for f in sorted(figs):
            if f <= LOCK_MAX_INT:  # no tocar 1..38
                continue
            nb = normalize_block(figs[f])
            stars = count_stars(nb)
            need_fix = (stars != 4) or (by_pat.get(pattern_tuple(nb), []).count(f) > 1)
            if need_fix:
                tries=20000
                while tries>0:
                    cand = random_block(ROWS_PER_FIG, cols_count, stars=4)
                    pat = pattern_tuple(cand)
                    if pat not in lock_patterns:
                        lock_patterns.add(pat)
                        result[f] = cand
                        break
                    tries -= 1
                if tries==0:
                    log.append(f"❗ No pude asignar patrón único a figura {f}; se deja como está.")

    # Reporte
    with open(args.report, "w", encoding="utf-8") as fp:
        fp.write("REPORTE DE AUDITORÍA DE FIGURAS\n")
        fp.write(f"Archivo  : {args.input}\n")
        fp.write(f"Salida   : {args.output}\n")
        fp.write(f"Filas/fig: {ROWS_PER_FIG}, columnas: {len(cols_idx)} (b..g)\n")
        fp.write(f"Total figs detectadas: {len(figs)} (esperadas ~{TOTAL_FIGS_EXPECTED})\n\n")
        if issues==0 and not dups:
            fp.write("✔ Sin problemas relevantes.\n\n")
        else:
            fp.write("HALLAZGOS:\n")
            for line in log:
                fp.write(line + "\n")

    # Escribir salida (si --fix, con correcciones; si no, igual vuelco ordenado)
    write_output(args.output, first_col, result, len(cols_idx))

    print("✅ Auditoría completada.")
    if args.fix:
        print("   Se aplicaron correcciones a figuras > 38 (si eran necesarias).")
    print(f"   Archivo salida: {args.output}")
    print(f"   Reporte       : {args.report}")
    if issues or dups:
        print("   Revisa el reporte para ver detalles por figura.")

if __name__ == "__main__":
    main()
