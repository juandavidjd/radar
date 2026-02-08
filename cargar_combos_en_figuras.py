#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
cargar_combos_en_figuras.py (versión robusta)
- Lee figuras desde XLSX (todo_fig_100_validado.xlsx).
- Lee combinaciones desde CSV/XLSX (todo_sum.csv o .xlsx).
- Encabezados tolerantes: 'combinacion' / 'combinación', 'signo1'/'signo 1'/'signo_1', etc.
- Autodetecta delimitador en CSV.
- Inserta los 4 dígitos de cada combinación en las posiciones '*' de cada figura.
- Guarda:
   • todo_fig_numeros.xlsx
   • todo_map.xlsx / todo_map.csv
"""

import sys, csv, argparse, random
from pathlib import Path

try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Alignment, Font, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("❌ Falta 'openpyxl'. Instala: pip install openpyxl")
    sys.exit(1)

# ---------- Parámetros ----------
FIG_XLSX_IN   = "todo_fig_100_validado.xlsx"
SUM_FILE_IN   = "todo_sum.csv"          # también puede ser .xlsx
FIG_XLSX_OUT  = "todo_fig_numeros.xlsx"
MAP_XLSX_OUT  = "todo_map.xlsx"
MAP_CSV_OUT   = "todo_map.csv"

ROWS_PER_FIG  = 4
GRID_LABELS   = ["b","c","d","e","f","g"]

# ---------- Helpers ----------
def _norm(s: str) -> str:
    """Normaliza encabezados: a minúsculas, sin espacios/guiones/bajos, sin acentos sencillos."""
    if s is None: return ""
    t = str(s).strip().lower()
    rep = {
        "á":"a","é":"e","í":"i","ó":"o","ú":"u","ñ":"n",
        "ä":"a","ë":"e","ï":"i","ö":"o","ü":"u",
    }
    t = "".join(rep.get(ch, ch) for ch in t)
    for ch in (" ", "\t", "_", "-", ".", "/"):
        t = t.replace(ch, "")
    return t

def read_figs_xlsx(path: Path):
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    headers=[]
    for c in range(1, ws.max_column+1):
        v = ws.cell(1,c).value
        headers.append("" if v is None else str(v).strip())
    name_to_idx = {headers[i].strip().lower(): i for i in range(len(headers))}
    cols_idx=[]
    for lab in GRID_LABELS:
        if lab.lower() in name_to_idx:
            cols_idx.append(name_to_idx[lab.lower()])
    if len(cols_idx) != 6:
        cols_idx = [i for i in range(1, min(7, len(headers)))]  # fallback

    figs = {}
    current = None
    for r in range(2, ws.max_row+1):
        sid = ws.cell(r,1).value
        sid = "" if sid is None else str(sid).strip()
        if sid.isdigit():
            current = int(sid)
            figs.setdefault(current, [])
        row_vals=[]
        for ci in cols_idx:
            v = ws.cell(r, ci+1).value
            row_vals.append("" if v is None else str(v).strip())
        if any(v != "" for v in row_vals):
            figs[current].append(row_vals)

    clean={}
    for f, blk in figs.items():
        b=[r[:] for r in blk]
        while b and all(c=="" for c in b[-1]): b.pop()
        if not b: continue
        if len(b) >= ROWS_PER_FIG: b=b[:ROWS_PER_FIG]
        else: b += [[""]*len(cols_idx) for _ in range(ROWS_PER_FIG-len(b))]
        clean[f]=b
    return wb, ws, headers, cols_idx, clean

def _sniff_delimiter(sample: str) -> str:
    # intenta detectar ; , \t
    if sample.count(";") > sample.count(",") and sample.count(";") >= sample.count("\t"):
        return ";"
    if sample.count("\t") > sample.count(","):
        return "\t"
    return ","  # por defecto

def read_sum_any(path: Path):
    """Devuelve lista de tuplas (combinacion, signo1, signo2, signo3). Soporta CSV o XLSX."""
    if not path.exists():
        return []
    if path.suffix.lower()==".xlsx":
        # Leer con openpyxl
        wb = load_workbook(path, data_only=True)
        ws = wb.active
        # encabezados normalizados
        raw_headers = []
        for c in range(1, ws.max_column+1):
            v = ws.cell(1, c).value
            raw_headers.append("" if v is None else str(v).strip())
        norm_headers = [_norm(h) for h in raw_headers]
        idx = {norm_headers[i]: i for i in range(len(norm_headers))}
        # posibles nombres
        key_comb = None
        for k in ("combinacion","combinacion4cifras","combo","combinacionfinal","combinacionresultado","combinacion_"):
            if k in idx: key_comb = k; break
        if key_comb is None:
            # probar con 'combinación' acentuada (ya normalizada)
            key_comb = "combinacion" if "combinacion" in idx else None

        def pick(*cands):
            for c in cands:
                if c in idx: return idx[c]
            return None

        i_comb = idx.get(key_comb) if key_comb else None
        i_s1 = pick("signo1","signo01","signo","s1","signo_1","signo 1")
        i_s2 = pick("signo2","signo02","s2","signo_2","signo 2")
        i_s3 = pick("signo3","signo03","s3","signo_3","signo 3")

        combos=[]
        if i_comb is not None:
            for r in range(2, ws.max_row+1):
                c = ws.cell(r, i_comb+1).value
                if c is None: continue
                comb = str(c).strip()
                if comb == "": continue
                s1 = str(ws.cell(r, (i_s1+1) if i_s1 is not None else 0).value).strip() if i_s1 is not None else ""
                s2 = str(ws.cell(r, (i_s2+1) if i_s2 is not None else 0).value).strip() if i_s2 is not None else ""
                s3 = str(ws.cell(r, (i_s3+1) if i_s3 is not None else 0).value).strip() if i_s3 is not None else ""
                combos.append((comb, s1 if s1!="None" else "", s2 if s2!="None" else "", s3 if s3!="None" else ""))
            return combos

        # si no hay columna obvia, probar 1ra columna como combinacion
        combos=[]
        for r in range(2, ws.max_row+1):
            c = ws.cell(r,1).value
            if c is None: continue
            comb=str(c).strip()
            if comb: combos.append((comb,"","",""))
        return combos

    # CSV
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        sample = f.read(2048)
    delim = _sniff_delimiter(sample)
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        rdr = csv.reader(f, delimiter=delim)
        rows = [ [ (x or "").strip() for x in row ] for row in rdr ]
    if not rows: return []

    headers = rows[0]
    data = rows[1:]
    norm = [_norm(h) for h in headers]
    idx = {norm[i]: i for i in range(len(norm))}

    # detectar columna combinación
    key_comb = None
    for k in ("combinacion","combinacion4cifras","combo","combinacionfinal","combinacionresultado","combinacion_"):
        if k in idx: key_comb = k; break
    if key_comb is None and "combinacion" in idx:
        key_comb = "combinacion"

    def pick(*cands):
        for c in cands:
            if c in idx: return idx[c]
        return None

    i_comb = idx.get(key_comb) if key_comb else None
    i_s1 = pick("signo1","signo01","signo","s1","signo_1","signo1","signo 1")
    i_s2 = pick("signo2","signo02","s2","signo_2","signo2","signo 2")
    i_s3 = pick("signo3","signo03","s3","signo_3","signo3","signo 3")

    combos=[]
    if i_comb is not None:
        for row in data:
            comb = row[i_comb] if i_comb < len(row) else ""
            if comb == "": continue
            s1 = row[i_s1] if (i_s1 is not None and i_s1 < len(row)) else ""
            s2 = row[i_s2] if (i_s2 is not None and i_s2 < len(row)) else ""
            s3 = row[i_s3] if (i_s3 is not None and i_s3 < len(row)) else ""
            combos.append((comb, s1, s2, s3))
        if combos:
            return combos

    # fallback: si hay una sola columna con la combinación
    if len(headers) >= 1:
        for row in data:
            if row and row[0]:
                combos.append((row[0], "", "", ""))
    return combos

def place_digits_in_block(block, combo_digits):
    out=[r[:] for r in block]
    coords=[]
    for i in range(len(out)):
        for j in range(len(out[0])):
            if out[i][j] == "*":
                coords.append((i,j))
    for k,d in enumerate(combo_digits[:4]):
        if k < len(coords):
            i,j = coords[k]
            out[i][j] = d
    return out

def save_figs_xlsx(figs_dict, headers, cols_idx, path_out: Path):
    wb=Workbook(); ws=wb.active; ws.title=path_out.stem
    first_header = headers[0] if headers and headers[0] else "fig"
    out_headers = [first_header] + GRID_LABELS[:len(cols_idx)]
    for j,h in enumerate(out_headers, start=1):
        ws.cell(1,j).value = h

    font=Font(name="Calibri", size=9)
    center=Alignment(horizontal="center", vertical="center")
    thin=Side(style="thin", color="DDDDDD")
    border=Border(left=thin,right=thin,top=thin,bottom=thin)

    r=2
    for fig in sorted(figs_dict):
        blk = figs_dict[fig]
        for i in range(ROWS_PER_FIG):
            ws.cell(r+i, 1).value = fig if i==0 else ""
            for j,val in enumerate(blk[i], start=2):
                cell=ws.cell(r+i, j)
                cell.value=val
                cell.font=font; cell.alignment=center; cell.border=border
            ws.cell(r+i,1).font=font; ws.cell(r+i,1).alignment=center
        r += ROWS_PER_FIG

    for j in range(1,8):
        ws.column_dimensions[get_column_letter(j)].width = 6 if j==1 else 5
    for i in range(2,r):
        ws.row_dimensions[i].height = 14

    wb.save(path_out)

def write_map(map_rows, xlsx_out: Path, csv_out: Path):
    wb=Workbook(); ws=wb.active; ws.title="map"
    headers=["fig","combinacion","signo1","signo2","signo3"]
    for j,h in enumerate(headers, start=1): ws.cell(1,j).value=h
    for i,row in enumerate(map_rows, start=2):
        for j,val in enumerate(row, start=1):
            ws.cell(i,j).value = val
    wb.save(xlsx_out)

    with csv_out.open("w", newline="", encoding="utf-8-sig") as f:
        w=csv.writer(f)
        w.writerow(headers)
        for row in map_rows:
            w.writerow(row)

def main():
    ap = argparse.ArgumentParser(description="Cargar combinaciones en figuras.")
    ap.add_argument("--fig", default=FIG_XLSX_IN, help="XLSX de figuras base (con '*').")
    ap.add_argument("--sum", default=SUM_FILE_IN, help="CSV/XLSX de combinaciones.")
    ap.add_argument("--out", default=FIG_XLSX_OUT, help="XLSX de salida con números.")
    ap.add_argument("--map-xlsx", default=MAP_XLSX_OUT, help="Mapa Excel figura↔combinación.")
    ap.add_argument("--map-csv", default=MAP_CSV_OUT, help="Mapa CSV figura↔combinación.")
    ap.add_argument("--aleatorio", action="store_true", help="Asignación aleatoria de combinaciones a figuras.")
    args = ap.parse_args()

    fig_path = Path(args.fig)
    sum_path = Path(args.sum)
    if not fig_path.exists():
        print(f"❌ No existe: {fig_path}"); sys.exit(1)
    if not sum_path.exists():
        print(f"❌ No existe: {sum_path}"); sys.exit(1)

    # Leer figuras
    wb, ws, headers, cols_idx, figs = read_figs_xlsx(fig_path)
    figs_ids = sorted(figs.keys())
    print(f"🧩 Figuras detectadas: {len(figs_ids)}")

    # Leer combinaciones
    combos = read_sum_any(sum_path)
    print(f"🔢 Combinaciones detectadas: {len(combos)}")
    if combos[:3]:
        print("   Ejemplos:", combos[:3])
    if not combos:
        print("⚠️ No se detectaron combinaciones en el archivo de sumas. Revisa encabezados y separador.")
        sys.exit(1)

    # Normalizar combinaciones a 4 dígitos
    def split_digits(c):
        digs = [ch for ch in c if ch.isdigit()]
        return (digs + ["0","0","0","0"])[:4]

    # Asignación
    pairs = list(zip(figs_ids, combos))
    if args.aleatorio:
        random.shuffle(combos)
        pairs = list(zip(figs_ids, combos))

    fig_num_blocks = {}
    map_rows = []
    for fig, tpl in pairs:
        comb, s1, s2, s3 = tpl
        digs = split_digits(comb)
        new_blk = place_digits_in_block(figs[fig], digs)
        fig_num_blocks[fig] = new_blk
        map_rows.append([fig, comb, s1, s2, s3])

    # Guardar
    save_figs_xlsx(fig_num_blocks, headers, cols_idx, Path(args.out))
    write_map(map_rows, Path(args.map_xlsx), Path(args.map_csv))

    print("✅ Proceso completo.")
    print(f"   Figuras con números: {args.out}")
    print(f"   Mapa figura-combinación: {args.map_xlsx} / {args.map_csv}")
    print(f"   Asignación: {'aleatoria' if args.aleatorio else '1→1, 2→2, …'}")

if __name__ == "__main__":
    main()
